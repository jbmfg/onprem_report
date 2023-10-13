import json
import trino
import xlsxwriter
import dateparser
import os
import openpyxl
from collections import defaultdict
from sqlite_connector import sqlite_db
from tesseract_connector import tesseract_connection
from datetime import datetime

class report_data(object):

    def __init__(self):
        self.customers = {}
        self.nulls = defaultdict(list)
        self.sfdb = tesseract_connection()
        self.db = sqlite_db("onprem_products.db")
        self.inst_ids = self.get_initial_list()
        self.act_dict = self.get_account_translation()

    def get_initial_list(self):
        query = f"""
        select i.installation_18_digit_id__c
        from edw_tesseract.sbu_ref_sbusfdc.installation__c i
        left join edw_tesseract.sbu_ref_sbusfdc.account a on i.account__c = a.id
        where  1=1
        --and a.cs_tier__c in ('Low', 'Medium', 'High', 'Holding')
        and i.product_group__c in ('Cb Protection', 'Cb Response', 'Cb Response Cloud')
        and i.installation_type__c in ('Perpetual', 'Subscription')
        and i.install_type__c in ('Partner', 'MSSP - Cb Protection', 'IR - Carbon Black', 'Other', 'General Availability', 'Bit9 Deployment', 'Initial purchase')
        and (i.cb_cloud_status__c not in ('Destroyed', 'Shutdown') or i.cb_cloud_status__c is null)
        and (i.status__c in ('New', 'In-Progress', 'Complete') or i.status__c is null)
        """
        query = """
        select distinct i.installation_18_digit_id__c
        from edw_tesseract.sbu_ref_sbusfdc.installation__c i
        left join edw_tesseract.sbu_ref_sbusfdc.bit9_subscriptions__c s on i.account__c = s.account__c
        where  1=1
        and s.active_subscription__c = True
        and s.product_group__c in ('Cb Response Cloud')
        """
        data = self.sfdb.execute(query)
        accts = ('0010h00001Znvh6AAB', '0013400001SaPxVAAV', '0018a00001kw5hxAAA', '0010h00001aAfkGAAS', '0010h00001ZmktrAAB', '0013400001QSgD4AAL', '0013400001P0HwuAAF', '0010h00001ZnFEbAAN', '0013400001LOZkpAAH', '0013400001OztdXAAR', '0010h00001ZxoPlAAJ', '0010h00001ZmwRpAAJ', '0018a00001kvSyyAAE', '0013400001NdRITAA3', '0010h00001cxYNfAAM', '0013400001TGxTkAAL', '0010h00001azAyIAAU', '0010h00001Ys8kpAAB', '0010h00001jTGDsAAO', '0013400001UaofUAAR', '0010h00001jTxn6AAC', '0010h00001ZmiR5AAJ',
 '0010h00001Zml2SAAR', '0010h00001dwigpAAA', '0013400001UOVktAAH', '0013400001S2lO8AAJ', '0010h00001ktg8NAAQ', '0013400001Rpc17AAB', '0013400001UqG0pAAF', '0010h00001Zn4SvAAJ', '0018000001CxkZjAAJ', '0013400001Pj8oRAAR', '0010h00001duwPBAAY', '0010h00001ZmumFAAR', '0018000001IncHMAAZ', '0010h00001Ys9iBAAR', '0010h00001k8bRtAAI', '0010h00001aAzkDAAS', '0018000001InODdAAN', '00180000014eJ43AAE', '0013400001QyqelAAB', '0013400001V04e9AAB', '0013400001TFGFjAAP', '0010h00001jUxEQAA0',
 '0010h00001cDXpRAAW', '0013400001MCE1LAAX', '0018a00001kvoGNAAY', '0010h00001jVPLpAAO', '0010h00001ZmyomAAB', '0010h00001cyd5tAAA', '0013400001RS5O5AAL', '0010h00001Zo9VXAAZ', '0010h00001YplmqAAB', '0018a00001mrv9DAAQ', '0013400001VSfGjAAL', '0013400001S0PH6AAN', '0010h00001YqcxOAAR', '0013400001WftwoAAB', '0010h00001ZmsOCAAZ', '0013000000D83bCAAR', '0010h00001dvcDRAAY', '0018a00001kvuf4AAA', '0013400001W6kPIAAZ', '0013400001T5oQ5AAJ', '0018000000rgdUJAAY', '0018000000h1vqAAAQ',
 '0010h00001ayuCDAAY', '0013400001T5k7wAAB', '0018000001FZEK7AAP', '0013400001WGrzHAAT', '0013400001LN9bZAAT', '0013400001V9DEdAAN', '0013000000DBENbAAP', '0010h00001Zx2CRAAZ', '0018a00001lamYPAAY', '0010h00001Zmm8KAAR', '0018a00001nAObCAAW', '0010h00001ZnGlOAAV', '0010h00001ZnK24AAF', '0013400001OZ12ZAAT', '0010h00001ayKhHAAU', '0018000000vcOubAAE', '0013000000D7sLaAAJ', '0013400001QUDReAAP', '0013400001M20mLAAR', '0013000000DBEI5AAP', '0010h00001YtKvvAAF', '0013000000FIm9TAAT',
 '0013400001TE7MQAA1', '0013400001K1KNvAAN', '0010h00001Ysny6AAB', '0010h00001iyGj1AAE', '0010h00001avpS9AAI', '0010h00001fdaE2AAI', '0010h00001dxncbAAA', '0010h00001YbafGAAR', '0013400001JVVBQAA5', '0013400001UcogtAAB', '0013400001W56lfAAB', '0010h00001iM0ZtAAK', '0013400001K2YnzAAF', '0013000000JgbzPAAR', '0010h00001ZnIlCAAV', '0013400001K234wAAB', '0010h00001a9L4QAAU', '0018000000wZ1pAAAS', '0018a00002GKS64AAH', '0018a000026xSi4AAE', '0010h00001aAgGaAAK', '0010h00001j6W7YAAU',
 '0013400001T5nf3AAB', '0013000000LW1RSAA1', '0010h00001cDUllAAG', '0010h00001ZmzfhAAB', '0010h00001iihGzAAI', '0013000000HISG0AAP', '0013400001SwAOtAAN', '0013400001MuK8eAAF', '0018000000rfgEzAAI', '0010h00001ktycuAAA', '0010h00001f3qNkAAI', '0010h00001Zmlf9AAB', '0010h00001ayB38AAE', '0010h00001ZxJboAAF', '0010h00001dwAq8AAE', '0010h00001fcKxTAAU', '0018a00001mnGOfAAM', '0018000000MfM82AAF', '0010h00001jSoBQAA0', '0010h00001ktCkcAAE', '0013400001K2t2pAAB', '0013400001RQVu8AAH',
 '0013000000EOPEZAA5', '0018000001HFPe6AAH', '0013000000D8jW6AAJ', '0010h00001YpdiLAAR', '0013400001UODidAAH', '0013400001WfonkAAB', '0013400001TEvFrAAL', '0010h00001jUoKdAAK', '0013400001WOyBKAA1', '0010h00001duFjCAAU', '0013000000E1Xk6AAF', '0010h00001ktFr7AAE', '0010h00001YqMiLAAV', '0013400001VehCSAAZ', '0013400001K0w5BAAR', '0018a00001w7AkmAAE', '0010h00001azQX1AAM', '0013400001LOHiEAAX', '0010h00001Ysl7LAAR', '0013400001QUhRlAAL', '0013000000DBEJ7AAP', '0013400001LO7hxAAD',
 '0018000001BVw48AAD', '0013400001NdIsTAAV', '0013400001RS9zaAAD', '0013400001TGMTLAA5', '0018000001GMaxpAAD', '0018000000Lixr7AAB', '0010h00001kt5NiAAI', '0010h00001hZ6ScAAK', '0013400001QUVsJAAX', '00180000018r7TyAAI', '0010h00001ikBPhAAM', '0013000000E1XgFAAV', '0018a00001upWJ5AAM', '0010h00001ZmYroAAF', '0010h00001ayDhmAAE', '0010h00001cDnpVAAS', '0010h00001ZmmDzAAJ', '0013400001NcjL3AAJ', '0013400001UMqZFAA1', '0013400001VecpCAAR', '0010h00001Zn1HlAAJ', '0013400001UsbniAAB',
 '0013400001TGWsdAAH', '0010h00001Zn492AAB', '0013400001UNv8lAAD', '0010h00001ksleMAAQ', '0010h00001cDeRjAAK', '0013400001QSS4GAAX', '0010h00001cxVF4AAM', '0010h00001ZmTxpAAF', '0013400001OxdQpAAJ', '0010h00001k8tRYAAY', '0010h00001duFuZAAU', '0010h00001ZoPRyAAN', '0013000000FIm9mAAD', '0010h00001ZofwMAAR', '0010h00001ZmsDHAAZ', '0013400001NcgdZAAR', '0010h00001k8i4oAAA', '0013400001Ubkh7AAB', '0013000000DT2ESAA1', '0013400001L1CNuAAN', '0018a00001n7e3oAAA', '0018a00001kuvInAAI',
 '0013000000D7sm8AAB', '0010h00001ZxZiWAAV', '0010h00001dwiiCAAQ', '0010h00001YsjVFAAZ', '0010h00001fcvaCAAQ', '0013400001TEzHTAA1', '0010h00001fdMyQAAU', '0010h00001dwA2GAAU', '0010h00001a9ntYAAQ', '0013000000D7scPAAR', '0010h00001aAAesAAG', '0013400001RTNiuAAH', '0013400001MD7LoAAL', '0018000001IqN4vAAF', '0010h00001YrjGSAAZ', '0010h00001aAiGkAAK', '0013400001T5s8cAAB', '0013400001MCnBfAAL', '0013400001KGnGqAAL', '0018000000p5Mj9AAE', '0010h00001cDpRsAAK', '0018000001HzuoOAAR',
 '0013400001UcYMGAA3', '0013400001T5oFgAAJ', '0013400001VhAv4AAF', '0018000000OOqjbAAD', '0013400001T5FMgAAN', '0010h00001YsfOxAAJ', '0010h00001f3vP5AAI', '0013400001RrEvzAAF', '0013400001M1PBtAAN', '0013000000D8jIZAAZ', '0010h00001ZnG01AAF', '0010h00001d1JCtAAM', '0010h00001ik0F0AAI', '0010h00001azVnvAAE', '0010h00001ik1xTAAQ', '0018a00001kvYhLAAU', '0013400001VV8HjAAL', '0013400001T5FaJAAV', '0010h00001Zn1RdAAJ', '0018000001BxCSEAA3', '0013400001OYjqUAAT', '0010h00001ikBrJAAU',
 '0013000000DBEIPAA5', '0013400001WNeGFAA1', '0018a00001upacwAAA', '0010h00001YseA8AAJ', '0010h00001cyxwEAAQ', '0013400001K2QO2AAN', '0010h00001cy788AAA', '0018a00001mEbxCAAS', '0010h00001Zn95OAAR', '0013400001SYQxIAAX', '0013400001THHQOAA5', '0013400001P0YZCAA3', '0013000000D8iBRAAZ', '0013400001SwOKgAAN', '0010h00001ZnHd1AAF', '0013400001TGHQyAAP', '0010h00001YrlwJAAR', '0013400001R2UXKAA3', '0018a00001zewUiAAI', '0013400001RQoqgAAD', '0013400001V0x3HAAR', '0010h00001Zy1qlAAB',
 '0013000000D8jEsAAJ', '0013400001Nbfd8AAB', '0010h00001axmUJAAY', '0013400001UMxo2AAD', '0018000000udyYCAAY', '0013000000EN7xkAAD', '0013000000EOPGrAAP', '0013400001JYUMzAAP', '0013400001RProGAAT', '0013400001NcODYAA3', '0013400001P0LTaAAN', '0010h00001ZnGrEAAV', '0010h00001d0xvuAAA', '0013400001S0yA2AAJ', '0013400001TFFEfAAP', '0013400001PhwtGAAR', '0010h00001jSvnDAAS', '0010h00001ZnKxTAAV', '0010h00001YpYrQAAV', '0010h00001ZnHm8AAF', '0010h00001YpdagAAB', '0010h00001ktcfOAAQ',
 '0013400001PjDN6AAN', '0010h00001j6p26AAA', '0010h00001YqWPjAAN', '0010h00001aA4eMAAS', '0013400001K0p1GAAR', '0013400001LOuk6AAD', '0013400001T5sG7AAJ', '0013400001RFUl7AAH', '0013400001NbrJgAAJ', '0010h00001ksnxOAAQ', '0013400001QRNM9AAP', '0013400001UMXtzAAH', '0010h00001dx1P0AAI', '0010h00001ZnJzcAAF', '0010h00001jUTdRAAW', '0013000000KwZF1AAN', '0013000000DT1AnAAL', '0018000000koKF0AAM', '00180000010jjbZAAQ', '0013400001SwE3FAAV', '0018000001IWI0uAAH', '0013400001MuaKkAAJ',
 '0013400001OwxDyAAJ', '0013400001T5jwFAAR', '0013400001OFz0uAAD', '0013000000D8iunAAB', '0010h00001Zn6vBAAR', '0013400001LOGQJAA5', '0013400001S3LB5AAN', '0013400001T5GPSAA3', '0013400001M1xwVAAR', '0018a000023dcIxAAI', '0013400001TFcdwAAD', '0013400001Wg4AnAAJ', '0013000000DBEDMAA5', '0010h00001cy16gAAA', '0010h00001ZxIJZAA3', '0018000000xRzTKAA0', '0010h00001dureqAAA', '0018a00001kvLuqAAE', '0010h00001hKcwpAAC', '0013400001LNvkTAAT', '0013400001SZlH7AAL', '0018000001G3RM7AAN',
 '0010h00001YrNhPAAV', '0013400001Sw6DaAAJ', '0013400001JWjr1AAD', '0018a00001ze5T8AAI', '0010h00001iwXz9AAE', '0010h00001axD2OAAU', '0013400001THK5kAAH', '0013400001MuMnCAAV', '0010h00001ZnEyLAAV', '0010h00001ZnJdVAAV', '0010h00001Ys1STAAZ', '0013400001L0yyMAAR', '0013000000D8jABAAZ', '0010h00001aAo3ZAAS', '0013400001JVB5zAAH')
        accts = "'" + "', '".join(accts) + "'"
        query = f"""
        select distinct i.installation_18_digit_id__c
        from edw_tesseract.sbu_ref_sbusfdc.installation__c i 
        where 1=1
        and i.product_group__c like '%Cb Response Cloud%'
        and i.account__c in ({accts})
        """

        # Insert into db just the inst_ids
        fields = ["inst_id"]
        self.db.insert("installations", fields, data)

        # Return a string that can be used in subsequent queries
        data = [i[0] for i in data]
        data = "'" + "', '".join(data) + "'"
        return data

    def get_installation_info(self):
        query = f"""
        select i.installation_18_digit_id__c,
        i.licenses_purchased__c,
        i.normalized_host_count__c,
        i.last_contact__c,
        i.account__c,
        i.product_group__c,
        i.sid__c,
        i.monitor_count__c,
        i.block_ask_count__c,
        i.lockdown_count__c,
        LOWER(i.carbon_black_alias__c),
        mp.name
        from edw_tesseract.sbu_ref_sbusfdc.installation__c i
        left join edw_tesseract.sbu_ref_sbusfdc.account mp on i.monitoring_partner__c = mp.account_id_18_digits__c
        where i.installation_18_digit_id__c in ({self.inst_ids})
        """
        data = self.sfdb.execute(query)
        fields = ("inst_id", "licenses_purchased", "normalized_host_count", "last_contact", "acct_id", "product",\
                 "sid", "le", "me", "he", "cb_alias", "monitoring_partner")
        self.db.update("installations", fields, data)

    def get_account_translation(self):
        query = f"""
        select i.account__c, i.id from
        edw_tesseract.sbu_ref_sbusfdc.installation__c i
        where i.installation_18_digit_id__c in ({self.inst_ids})
        and i.account__c is not Null
        """
        data = self.sfdb.execute(query)
        act_dict = defaultdict(list)
        for act_id, inst_id in data:
            act_dict[act_id].append(inst_id)
        return act_dict

    def get_account_info(self):
        accts = "'" + "', '".join(self.act_dict.keys()) + "'"
        query = f"""
        select
        a.account_id_18_digits__c,
        a.cs_tier__c,
        a.arr__c,
        a.name,
        a.GS_CSM_Meter_Score__c,
        a.csm_meter_comments__c,
        a.GS_Overall_Score__c,
        a.gs_adoption_comments__c,
        csm.name,
        man.name,
        cse.name,
        a.owner_name__c,
        a.vmstar_geo__c,
        a.vmstar_sub_division__c,
        a.vmstar_cm_country__c,
        csp.name
        from edw_tesseract.sbu_ref_sbusfdc.account a
        left join edw_tesseract.sbu_ref_sbusfdc.user_sbu csm on a.Assigned_CP__c = csm.Id
        left join edw_tesseract.sbu_ref_sbusfdc.user_sbu man on csm.managerid = man.id
        left join edw_tesseract.sbu_ref_sbusfdc.user_sbu cse on a.Customer_Success_Engineer__c = cse.Id
        left join edw_tesseract.sbu_ref_sbusfdc.account csp on a.cs_partner__c = csp.account_id_18_digits__c
        where a.account_id_18_digits__c in ({accts})
        """
        data = self.sfdb.execute(query)
        fields = ["acct_id", "tier", "arr", "account_name", "csm_score", "csm_comments"]
        fields += ["gs_score", "adoption_comments", "csm", "csm_manager", "cse", "account_manager"]
        fields += ["vmw_geo", "vmw_sub_div", "vmw_country", "cs_partner"]
        self.db.insert("accounts", fields, data)

    def get_opportunity_info(self):
        accts = "'" + "', '".join(self.act_dict.keys()) + "'"
        query = f"""
        select o.id,
        o.accountid,
        o.acv_amount__c,
        o.cb_forecast__c,
        o.closeDate,
        o.product_family__c
        from edw_tesseract.sbu_ref_sbusfdc.opportunity o
        where o.accountid in ({accts})
        and o.closedate > CURRENT_DATE
        and o.type like '%Renewal%'
        """
        data = self.sfdb.execute(query)
        fields = ("opp_id", "acct_id", "acv", "forecast", "close_date", "type")
        self.db.insert("opportunities", fields, data)

    def get_subscription_info(self):
        accts = "'" + "', '".join(self.act_dict.keys()) + "'"
        query = f"""
        select account__c,
        coalesce(arr__c, 0.0),
        end_date__c,
        id,
        product_description__c,
        product__c,
        product_group__c,
        quantity__c,
        coalesce(subscription_term__c, 0),
        coalesce(tcv__c, 0.0)
        from edw_tesseract.sbu_ref_sbusfdc.bit9_subscriptions__c s
        where active_subscription__c = true
        and account__c in ({accts})
        """
        data = self.sfdb.execute(query)
        fields = ["acct_id", "arr", "end_date", "sub_id"]
        fields += ["description", "product_id", "product"]
        fields += ["quantity", "sub_term", "tcv"]
        self.db.insert("subscriptions", fields, data)

    def get_cta_info(self):
        accts = "'" + "', '".join(self.act_dict.keys()) + "'"
        for cta_type in ("Product Usage Analytics", "Tech Assessment", "CSA Whiteboarding"):
            fields = ("acct_id", "cta_type", "closed_date", "status")
            query = f"""
            select account_id,
            '{cta_type}',
            max(closed_date),
            case when status in ('New','Work In Progress') then 'Open' else 'Closed' end
            from edw_tesseract.sbu_ref_sbusfdc.gsctadataset
            where reason like '{cta_type}'
            and account_id in ({accts})
            and status not in ('Closed No Action', 'Closed Unsuccessful', 'Closed Invalid')
            group by account_id, status, closed_date
            """
            data = self.sfdb.execute(query)
            self.db.insert("ctas", fields, data)

    def renewal_quarter(self):
        def lookup_q(opp_date):
            formatstr = "%Y-%m-%d"
            if isinstance(opp_date, str):
                opp_date = datetime.strptime(opp_date, formatstr)
            opp_date = datetime(opp_date.year, opp_date.month, opp_date.day)
            qdict = {
                "2021": {
                    "Q1": ["2020-02-01", "2020-04-30"],
                    "Q2": ["2020-05-01", "2020-07-30"],
                    "Q3": ["2020-07-31", "2020-10-29"],
                    "Q4": ["2020-10-30", "2021-01-28"]
                },
                "2022": {
                    "Q1": ["2021-01-29", "2021-04-29"],
                    "Q2": ["2021-04-30", "2021-07-29"],
                    "Q3": ["2021-07-30", "2021-10-28"],
                    "Q4": ["2021-10-29", "2022-01-27"]
                },
                "2023": {
                    "Q1": ["2022-01-28", "2022-04-28"],
                    "Q2": ["2022-04-29", "2022-07-28"],
                    "Q3": ["2022-07-29", "2022-10-27"],
                    "Q4": ["2022-10-28", "2023-01-26"]
                },
                "2024": {
                    "Q1": ["2023-01-27", "2023-04-27"],
                    "Q2": ["2023-04-28", "2023-07-27"],
                    "Q3": ["2023-07-28", "2023-10-26"],
                    "Q4": ["2023-10-27", "2024-01-25"]
                },
                "2025": {
                    "Q1": ["2024-01-26", "2024-04-25"],
                    "Q2": ["2024-04-26", "2024-07-25"],
                    "Q3": ["2024-07-26", "2024-10-24"],
                    "Q4": ["2024-10-25", "2025-01-23"]
                },
                "2026": {
                    "Q1": ["2025-01-24", "2025-04-24"],
                    "Q2": ["2025-04-25", "2025-07-24"],
                    "Q3": ["2025-07-25", "2025-10-23"],
                    "Q4": ["2025-10-24", "2026-01-22"]}
                }
            opp_year = opp_date.year
            finding = "Unknown"
            for year in qdict:
                for q in qdict[year]:
                    start = datetime.strptime(qdict[year][q][0], formatstr)
                    end = datetime.strptime(qdict[year][q][1], formatstr)
                    if start <= opp_date <= end:
                        finding = f"{year} {q}"
            return finding

        query = "select opp_id, close_date from opportunities;"
        data = self.db.execute(query)
        data = [[i[0], lookup_q(i[1])] for i in data]
        fields = ("opp_id", "renewal_qt")
        self.db.update("opportunities", fields, data)

    def deployment_percentage(self):
        query = "select inst_id, normalized_host_count, licenses_purchased from installations;"
        data = self.db.execute(query)
        deployments = []
        for i in data:
            if i[1] == 0:
                deployment = "0%"
            elif i[1] is not None and i[2]:
                deployment = f"{round(i[1]/i[2] * 100, 2)}%"
            else:
                continue
            deployments.append([i[0], deployment])
        fields = ("inst_id", "deployment")
        self.db.update("installations", fields, deployments)

    def enforcement_levels(self):
        for el in ("le", "me", "he"):
            query = f"select inst_id, {el}, licenses_purchased from installations;"
            data = self.db.execute(query)
            percs = []
            for i in data:
                if i[1] == 0:
                    perc = "0%"
                elif i[1] is not None and i[2]:
                    perc = f"{round(i[1]/i[2] * 100, 2)}%"
                else:
                    continue
                percs.append([i[0], perc])
            fields = ("inst_id", f"{el}_perc")
            self.db.update("installations", fields, percs)

    def air_gapped(self):
        for product in ("Cb Protection", "Cb Response", "Cb Response Cloud"):
            query = f"""
            select
            i.inst_id,
            case
            when i.last_contact >
                DATE((select max(last_contact) from installations where product = '{product}'), '-5 Days')
                then False else True end
            from installations i
            where i.product = '{product}';
            """
            data = self.db.execute(query)
            fields = ("inst_id", "air_gapped")
            self.db.update("installations", fields, data)

    def product_family(self):
        query = "select distinct type from opportunities;"
        data = [i[0] for i in self.db.execute(query)]
        products = set([i for prods in data for i in prods.split(";")])

    def get_s3(self):
        xlsx_file = "HEDR Hosted S3 Buckets.xlsx"
        data = []
        wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        s = wb["Instances"]
        for x, i in enumerate(s.rows):
            alias = s.cell(row=x+1, column=1).value.lower().replace("-", "_")
            bucket = s.cell(row=x+1, column=2).value
            data.append([alias, bucket])
        fields = ("alias", "s3_bucket_name")
        self.db.insert("s3", fields, data)

    def get_activity(self):
        data = []
        for f in xlsx_files:
            wb = openpyxl.load_workbook(f, data_only=True)
            s = wb["Mda Sheet"]
            for x, i in enumerate(s.rows):
                account = s.cell(row=x+1, column=1).value
                act_date = s.cell(row=x+1, column=6).value
                act_date = dateparser.parse(act_date, settings={'TIMEZONE': 'UTC'})
                if not act_date:
                    continue
                act_date = datetime.strftime(act_date, "%Y-%m-%d")
                data.append([account, act_date])
        fields = ["acct_id", "activity_date"]
        self.db.insert("cse_activity", fields, data)

def table_creations():
    db = sqlite_db("onprem_products.db")
    for table in ("installations", "accounts", "opportunities", "subscriptions",\
                  "cse_activity", "ctas", "inst_summary", "acct_summary", "s3"):
        db.execute(f"drop table if exists {table};")

    # CSE Timeline Activities
    query = """
    CREATE table cse_activity(
    acct_id TEXT,
    activity_date TEXT
    );
    """
    db.execute(query)

    # Installations
    query = """
    CREATE table installations(
    inst_id TEXT PRIMARY KEY,
    licenses_purchased INTEGER DEFAULT Null CHECK (typeof(licenses_purchased) in ('integer', Null)),
    normalized_host_count INTEGER DEFAULT Null CHECK (typeof(normalized_host_count) in ('integer', Null)),
    deployment TEXT DEFAULT Null,
    last_contact STRING,
    acct_id STRING,
    product STRING,
    air_gapped INTEGER DEFAULT Null CHECK (typeof(air_gapped) in ('integer', Null)),
    sid STRING DEFAULT Null,
    le INTEGER DEFAULT 0 CHECK (typeof(le) in ('integer', Null)),
    le_perc TEXT DEFAULT NULL,
    me INTEGER DEFAULT 0 CHECK (typeof(me) in ('integer', Null)),
    me_perc TEXT DEFAULT NULL,
    he INTEGER DEFAULT 0 CHECK (typeof(he) in ('integer', Null)),
    he_perc TEXT DEFAULT NULL,
    cb_alias TEXT DEFAULT NULL,
    monitoring_partner TEXT DEFAULT NULL
    );
    """
    db.execute(query)

    # Accounts
    query = """
    CREATE table accounts(
    acct_id TEXT PRIMARY KEY,
    tier TEXT,
    arr INTEGER DEFAULT 0 CHECK (typeof(arr) in ('integer', Null)),
    account_name TEXT,
    csm_score INTEGER DEFAULT 0 CHECK (typeof(csm_score) in ('integer', Null)),
    csm_comments TEXT,
    gs_score INTEGER DEFAULT 0 CHECK (typeof(gs_score) in ('integer', Null)),
    adoption_comments TEXT,
    csm TEXT,
    csm_manager TEXT,
    cse TEXT,
    account_manager TEXT,
    vmw_geo TEXT,
    vmw_sub_div TEXT,
    vmw_country TEXT,
    cs_partner TEXT
    );
    """
    db.execute(query)

    # Opportunities
    query = """
    CREATE table opportunities(
    opp_id TEXT PRIMARY KEY,
    acct_id TEXT,
    acv INTEGER CHECK (typeof(acv) in ('integer', Null)),
    forecast TEXT,
    close_date TEXT,
    renewal_qt TEXT,
    type TEXT
    );
    """
    db.execute(query)

    # Subscriptions
    query = """
    CREATE table subscriptions(
    acct_id TEXT,
    arr REAL CHECK (typeof(arr) in ('real')),
    end_date TEXT,
    sub_id TEXT,
    description TEXT,
    product_id TEXT,
    product TEXT,
    quantity INTEGER CHECK (typeof(quantity) in ('integer')),
    sub_term INTEGER CHECK (typeof(sub_term) in ('integer')),
    tcv REAL CHECK (typeof(tcv) in ('real'))
    );
    """
    db.execute(query)

    # CTAs
    query = """
    CREATE table ctas(
    acct_id TEXT,
    cta_type TEXT,
    closed_date TEXT,
    status TEXT
    );
    """
    db.execute(query)

    # Installation Summary
    query = """
    CREATE table inst_summary (
    inst_id TEXT,
    sid TEXT,
    licenses_purchased INTEGER DEFAULT 0 CHECK (typeof(licenses_purchased) in ('integer', Null)),
    normalized_host_count INTEGER DEFAULT 0 CHECK (typeof(normalized_host_count) in ('integer', Null)),
    deployment REAL DEFAULT 0 CHECK (typeof(deployment) in ('REAL', Null)),
    le INTEGER DEFAULT 0 CHECK (typeof(le) in ('integer', Null)),
    le_perc TEXT DEFAULT NULL,
    me INTEGER DEFAULT 0 CHECK (typeof(me) in ('integer', Null)),
    me_perc TEXT DEFAULT NULL,
    he INTEGER DEFAULT 0 CHECK (typeof(he) in ('integer', Null)),
    he_perc TEXT DEFAULT NULL,
    last_contact TEXT,
    acct_id TEXT,
    product TEXT,
    air_gapped TEXT,
    tier TEXT,
    arr INTEGER DEFAULT 0 CHECK (typeof(arr) in ('integer', Null)),
    account_name TEXT,
    csm_score INTEGER DEFAULT 0 CHECK (typeof(csm_score) in ('integer', Null)),
    csm_comments TEXT,
    gs_score INTEGER DEFAULT 0 CHECK (typeof(gs_score) in ('integer', Null)),
    adoption_comments TEXT,
    csm TEXT,
    csm_manager TEXT,
    cse TEXT,
    close_date TEXT,
    renewal_qt TEXT,
    forecast TEXT,
    opp_acv INTEGER DEFAULT 0 CHECK (typeof(opp_acv) in ('integer', Null)),
    opp_count INTEGER DEFAULT 0 CHECK (typeof(opp_count) in ('integer', Null)),
    sub_product_arr INTEGER DEFAULT 0 CHECK (typeof(sub_product_arr) in ('integer', Null)),
    product_usage_analytics TEXT,
    tech_assessment TEXT,
    csa_whiteboarding TEXT,
    last_timeline TEXT,
    account_manager TEXT,
    vmw_geo TEXT,
    vmw_sub_div TEXT,
    vmw_country TEXT,
    monitoring_partner TEXT DEFAULT NULL,
    cb_alias TEXT DEFAULT NULL,
    cs_partner TEXT DEFAULT NULL);
    """
    db.execute(query)

    # account summary
    query = """
    CREATE TABLE acct_summary (
    acct_id TEXT,
    tier TEXT,
    arr INTEGER DEFAULT 0 CHECK (typeof(arr) in ('integer', Null)),
    account_name TEXT,
    csm_score INTEGER DEFAULT 0 CHECK (typeof(csm_score) in ('integer', Null)),
    csm_comments TEXT,
    gs_score INTEGER DEFAULT 0 CHECK (typeof(gs_score) in ('integer', Null)),
    adoption_comments TEXT,
    csm TEXT,
    csm_manager TEXT,
    cse TEXT,
    product TEXT,
    last_timeline TEXT,
    product_usage_analytics TEXT,
    tech_assessment TEXT,
    csa_whiteboarding TEXT,
    connected_count INTEGER DEFAULT 0 CHECK (typeof(connected_count) in ('integer', Null)),
    disconnected_count INTEGER DEFAULT 0 CHECK (typeof(disconnected_count) in ('integer', Null)),
    renewal_date TEXT,
    renewal_qt TEXT,
    forecast TEXT,
    product_acv INTEGER DEFAULT 0 CHECK (typeof(product_acv) in ('integer', Null)),
    licenses_purchased INTEGER DEFAULT 0 CHECK (typeof(licenses_purchased) in ('integer', Null)),
    sub_deployment_perc REAL DEFAULT 0 CHECK (typeof(sub_deployment_perc) in ('REAL', Null)),
    inst_deployment_perc REAL DEFAULT 0 CHECK (typeof(inst_deployment_perc) in ('REAL', Null)),
    le INTEGER DEFAULT 0 CHECK (typeof(le) in ('integer', Null)),
    le_perc TEXT DEFAULT NULL,
    me INTEGER DEFAULT 0 CHECK (typeof(me) in ('integer', Null)),
    me_perc TEXT DEFAULT NULL,
    he INTEGER DEFAULT 0 CHECK (typeof(he) in ('integer', Null)),
    he_perc TEXT DEFAULT NULL,
    products TEXT,
    account_manager TEXT,
    vmw_geo TEXT,
    vmw_sub_div TEXT,
    vmw_country TEXT,
    monitoring_partner TEXT DEFAULT NULL,
    cs_partner TEXT DEFAULT NULL,
    cb_alias TEXT DEFAULT NULL,
    s3_bucket INTEGER DEFAULT 0);
    """
    db.execute(query)

    # S3 buckets
    query = """
    CREATE TABLE s3 (
        alias TEXT DEFAULT NULL,
        s3_bucket_name TEXT DEFAULT NULL);
    """
    db.execute(query)

def writerows(wb, sheet, data, linkBool=False, setwid=True, col1url=False, bolder=False):
    bold = wb.add_format({"bold": True})
    # first get the length of the longest sting to set column widths
    numCols = len(data[0])
    widest = [10 for _ in range(numCols)]
    if setwid:
        try:
            for i in data:
                for x in range(len(data[0])):
                    if type(i[x]) == int:
                        pass
                    elif i[x] is None:
                        pass
                    elif not isinstance(i[x], float) and widest[x] < len(i[x].encode("ascii", "ignore")):
                        if len(str(i[x])) > 50:
                            widest[x] = 50
                        else:
                            widest[x] = len(str(i[x])) #+ 4
        except IndexError:
            pass
            # print ("--INFO-- Index Error when setting column widths")
        except TypeError:
            print ("type error")
        except AttributeError:
            # Added check for floats above so this probably isnt needed any more
            print ("\n--INFO-- Can't encode a float\n")

    for x, i in enumerate(widest):
        sheet.set_column(x, x, i)

    # Then write the data
    for r in data:
        for i in r:
            if type(i) == str:
                i = i.encode("ascii", "ignore")
    counter = 0
    for x, r in enumerate(data):
        counter += 1
        cell = "A" +str(counter)
        if bolder and (data[x-1] == "" or x==0):
            sheet.write_row(cell, r, bold)
        else:
            sheet.write_row(cell, r)
        if col1url:
            if x == 0:
                pass
            else:
                sheet_name = f"{x-1}. {r[0]}"[:31]
                sheet.write_url(cell, f"internal:'{sheet_name}'!A1", string=r[0])
        if linkBool:
            sheet.write_url(0, 6, "internal:Master!A1", string="Mastersheet")
    return True

def create_inst_master(db, prod):
    def add_metric(data_dict, new_data):
        new_keys = []
        for row in new_data:
            row_tup = list(zip(row.keys(), [i for i in row]))
            for i in row.keys(): new_keys.append(i)
            inst_id = row_tup.pop(0)[1]
            data_dict[inst_id].update(dict(row_tup))
        for inst_id in data_dict:
            for key in set(new_keys):
                if key not in data_dict[inst_id]:
                    # inst_id is the first level key, we dont need it in the value keys too
                    if key == "inst_id": continue
                    data_dict[inst_id][key] = None
        return data_dict

    rows = defaultdict(dict)

    # All of installations
    data = db.execute_dict(f"select * from installations where product = '{prod}';")
    add_metric(rows, data)

    # All of accounts
    query = f"""
    select i.inst_id, a.*
    from installations i
    left join accounts a on i.acct_id = a.acct_id
    where i.product = '{prod}';
    """
    data = db.execute_dict(query)
    add_metric(rows, data)

    # Those opportunities that apply *CBLO can be multiple so its omitted + wtf is other?
    # Provides metrics related only to the next renewal for the product in question
    lookup = {
        "Cb Cloud": ["CBWL", "CBVM", "CBWS", "CBD", "CBCO", "CBTS", "CBTH"],
        "Cb Response Cloud": ["CBRC"],
        "Cb Protection": ["CBP"],
        "Cb Response": ["CBR"]
    }
    query = f"""
    select i.inst_id,
    o.close_date,
    o.renewal_qt,
    o.forecast,
    o.acv as opp_acv,
    count(*) as opp_count
    from installations i
    left join opportunities o on i.acct_id = o.acct_id
    inner join
        (select opp_id,
        min(close_date) cd
        from opportunities
        group by opp_id ) o2
        on o.opp_id = o2.opp_id and o.close_date = o2.cd
    where i.product = '{prod}'
    and o.type like '%{", ".join(lookup[prod])}%'
    group by i.inst_id
    order by o.close_date desc;
    """
    data = db.execute_dict(query)
    add_metric(rows, data)

    # Arr from just the product in question
    query = f"""
    select i.inst_id,
    round(sum(s.arr), 2) sub_product_arr
    from installations i
    left join subscriptions s on i.acct_id = s.acct_id and i.product = s.product
    where 1=1
    and i.product = '{prod}'
    group by i.inst_id
    """
    data = db.execute_dict(query)
    add_metric(rows, data)

    # CTAs from gainsight
    for cta in ("Product Usage Analytics", "Tech Assessment", "CSA Whiteboarding"):
        query = f"""
        select i.inst_id,
        max(c.closed_date) as '{cta.lower().replace(" ", "_")}'
        from installations i
        left join ctas c on i.acct_id = c.acct_id
        where c.cta_type = '{cta}'
        and c.status = 'Closed'
        and i.product = '{prod}'
        group by i.inst_id
        """
        data = db.execute_dict(query)
        add_metric(rows, data)

    # CSE Timeline activities
    query = f"""
    select i.inst_id,
    max(cse.activity_date) as 'last_timeline'
    from installations i
    left join accounts a on i.acct_id = a.acct_id
    left join cse_activity cse on a.account_name = cse.acct_id
    where i.product = '{prod}'
    group by i.inst_id
    """
    data = db.execute_dict(query)
    add_metric(rows, data)

    fields = ["inst_id"] + list(rows[list(rows)[0]].keys())
    rows = [[inst_id] + list(rows[inst_id].values()) for inst_id in rows]
    db.insert("inst_summary", fields, rows)
    return rows

def create_acct_master(db, prod):
    def add_metric(data_dict, new_data):
        new_keys = []
        for row in new_data:
            row_tup = list(zip(row.keys(), [i for i in row]))
            for i in row.keys(): new_keys.append(i)
            acct_id = row_tup.pop(0)[1]
            if acct_id not in data_dict: continue
            data_dict[acct_id].update(dict(row_tup))
        for acct_id in data_dict:
            for key in set(new_keys):
                if key not in data_dict[acct_id]:
                    # acct_id is the first level key, we dont need it in the value keys too
                    if key == "acct_id": continue
                    data_dict[acct_id][key] = None
        return data_dict

    # Seed table with just the accounts that have the product in question
    data = [i[0] for i in db.execute(f"select acct_id from installations where product = '{prod}';")]
    rows = {i:{} for i in data}

    # All of accounts table
    data = db.execute_dict(f'select *, "{prod}" as product from accounts;')
    add_metric(rows, data)

    # CSE Timeline activities
    query = """
    select a.acct_id,
    max(cse.activity_date) as 'last_timeline'
    from accounts a
    left join cse_activity cse on a.account_name = cse.acct_id
    group by a.acct_id;
    """
    data = db.execute_dict(query)
    add_metric(rows, data)

    # Ctas
    for cta in ("Product Usage Analytics", "Tech Assessment", "CSA Whiteboarding"):
        query = f"""
        select a.acct_id,
        max(c.closed_date) as '{cta.lower().replace(" ", "_")}'
        from accounts a
        left join ctas c on a.acct_id = c.acct_id
        where c.cta_type = '{cta}'
        and c.status = 'Closed'
        group by a.acct_id;
        """
        data = db.execute_dict(query)
        add_metric(rows, data)

    # Deployment info from installations
    query = f"""
    select a.acct_id,
    sum(case when i.air_gapped = 0 then i.normalized_host_count end) as connected_count,
    sum(case when i.air_gapped = 1 then i.normalized_host_count end) as disconnected_count,
    group_concat(distinct i.monitoring_partner) as monitoring_partner,
    group_concat(distinct i.cb_alias) as cb_alias
    from accounts a
    left join installations i on a.acct_id = i.acct_id
    where i.product = '{prod}'
    group by a.acct_id;
    """
    data = db.execute_dict(query)
    add_metric(rows, data)

    # s3
    query = f"""
    select a.acct_id,
    case when s3.alias is Null then 0 else 1 end as s3_bucket
    from accounts a
    left join installations i on a.acct_id = i.acct_id
    left join s3 on i.cb_alias = s3.alias
    where i.product = '{prod}'
    group by a.acct_id;
    """
    data = db.execute_dict(query)
    add_metric(rows, data)
    print(query)

    # Opportunities
    lookup = {
        "Cb Cloud": ["CBWL", "CBVM", "CBWS", "CBD", "CBCO", "CBTS", "CBTH", "Endpoint STD", "EEDR", "Endpoint"],
        "Cb Response Cloud": ["Hosted EDR"],
        "Cb Protection": ["CBP", "Application Control"],
        "Cb Response": ["CBR"]
    }
    query = f"""
    select a.acct_id,
    group_concat(o.close_date) as renewal_date,
    group_concat(o.renewal_qt) as renewal_qt,
    group_concat(o.forecast) as forecast --,
    --sum(o.acv) as product_acv
    from accounts a
    left join opportunities o on a.acct_id = o.acct_id
    where o.type like '%{"%, %".join(lookup[prod])}%'
    group by a.acct_id;
    """
    data = db.execute_dict(query)
    add_metric(rows, data)

    # purchased licenses from subscriptions
    query = f"""
    select acct_id,
    sum(quantity) as licenses_purchased,
    sum(arr) as product_acv
    from subscriptions
    where product = '{prod}'
    group by acct_id;
    """
    data = db.execute_dict(query)
    add_metric(rows, data)

    # Calculated fields
    # Deployment percentage from subscriptions
    query = f"""
    select hc.acct_id,
    round(cast(nhc as real) / quan * 100, 2) as sub_deployment_perc
    from (
        select i.acct_id,
        sum(i.normalized_host_count) nhc
        from installations i
        where i.product = '{prod}'
        and i.air_gapped = 0
        group by i.acct_id) as hc
    join (
        select s.acct_id,
        sum(s.quantity) quan
        from subscriptions s
        where s.product = '{prod}'
        group by s.acct_id) as ss on hc.acct_id = ss.acct_id
    """
    data = db.execute_dict(query)
    add_metric(rows, data)

    # Deployment percentage by getting max from installation records
    query = f"""
    select hc.acct_id,
    round(cast(nhc as real) / quan * 100, 2) as inst_deployment_perc
    from (
        select i.acct_id,
        sum(i.normalized_host_count) nhc
        from installations i
        where i.product = '{prod}'
        group by i.acct_id) as hc
    join (
        select i.acct_id,
        max(i.licenses_purchased) quan
        from installations i
        where i.product = '{prod}'
        group by i.acct_id) as ss on hc.acct_id = ss.acct_id
    """
    data = db.execute_dict(query)
    add_metric(rows, data)

    # Enforcement Levels
    query = f"""
    select i.acct_id,
    sum(i.le) as le,
    round(cast(sum(i.le) as real) / max(i.licenses_purchased) * 100, 2) as le_perc,
    sum(i.me) as me,
    round(cast(sum(i.me) as real) / max(i.licenses_purchased) * 100, 2) as me_perc,
    sum(i.he) as he,
    round(cast(sum(i.he) as real) / max(i.licenses_purchased) * 100, 2) as he_perc
    from installations i
    where i.product = '{prod}'
    and i.air_gapped = 0
    group by i.acct_id;
    """
    data = db.execute_dict(query)
    add_metric(rows, data)

    # Products owned
    query = f"""
    select i.acct_id,
    GROUP_CONCAT(DISTINCT i.product) || "," || group_concat(DISTINCT s.product)
    from installations i
    left join subscriptions s on i.acct_id = s.acct_id
    where i.product = '{prod}'
    group by i.acct_id
    """
    data = [list(i) for i in db.execute(query)]
    replacements = (
        ("cb protection", "AC"),
        ("cb response", "EDR"),
        ("cb response cloud", "HEDR"),
        ("EDR cloud", "HEDR"),
        ("cb threathunter", "EEDR"),
        ("cb defense", "ES"),
        ("carbon black endpoint standard", "ES"),
        ("cb liveops", "Live Ops"),
        ("cb workload", "Workloads"),
        ("cb threatsight", "ThreatSight"),
        ("endpoint enterprise", "Endpoint Enterprise"),
        ("endpoint advanced", "Endpoint Advanced"),
        ("vmware workspace security", "Workspace Security"),
        ("carbon black ", "")
    )
    for row in data:
        if not row[1]: continue
        prods = list(set(row[1].lower().split(",")))
        for rpl in replacements:
            prods = [i.replace(rpl[0], rpl[1]) for i in prods]
            row[1] = ", ".join(prods)
    for acct_id in rows:
        rows[acct_id]["products"] = None
    for acct_id, products in data:
        if acct_id in rows:
            rows[acct_id]["products"] = products
    fields = ["acct_id"] + list(rows[list(rows)[0]].keys())
    new_rows = []
    for acct_id in rows:
        row = [acct_id]
        fields = ["acct_id"]
        for key in rows[acct_id]:
            fields.append(key)
            row.append(rows[acct_id][key])
        new_rows.append(row)

    rows = [[acct_id] + list(rows[acct_id].values()) for acct_id in rows]
    db.insert("acct_summary", fields, new_rows)

def write_report(db, product):
    lookup = {"Cb Response Cloud": "HEDR", "Cb Protection": "AC", "Cb Response": "EDR"}
    type_lookup = {"Cb Response Cloud": "cbrc", "Cb Protection": "cbp", "Cb Response": "cbr"}
    wb = xlsxwriter.Workbook(f"Consumption Report_{product}.xlsx")

    # Account Level
    sheet = wb.add_worksheet("Accounts")
    query = f"""
    select
    account_name as "Account",
    products as "Products Owned",
    renewal_date as "Next Renewal",
    renewal_qt as "Renewal Qt",
    forecast as "Renwewal Forecast",
    tier as "Tier",
    --monitoring_partner || ", " || cs_partner as "Partner",
    cs_partner as "Partner",
    csm as "CSM",
    csm_manager as "CSM Manager",
    cse as "CSE",
    account_manager as "Account Manager",
    vmw_geo as "VMW Geo",
    vmw_sub_div as "VMW Sub-division",
    vmw_country as "VMW Country",
    arr as "ARR",
    product_acv as "Product ACV",
    csm_score as "CSM Score",
    gs_score as "GS Score",
    csm_comments as "CSM Comments",
    adoption_comments as "Adoption Comments",
    last_timeline as "Latest CSE Activity",
    product_usage_analytics as "Last CUA",
    tech_assessment as "Last TA",
    csa_whiteboarding as "Last WB",
    connected_count as "Normalized Endpoints",
    disconnected_count as "Disconnected Endpoints",
    licenses_purchased as "Licenses",
    le as "LE Count",
    le_perc as "LE Perc",
    me as "ME Count",
    me_perc as "ME Perc",
    he as "HE Count",
    he_perc as "HE Perc",
    sub_deployment_perc as "Deployment(Sub)",
    inst_deployment_perc as "Deployment(Inst)",
    s3_bucket as "Have S3 Bucket",
    acct_id as "Account ID"
    from acct_summary
    where product = '{prod}'
    order by account_name;
    """
    data = db.execute_dict(query)
    
    # Clean up data that doesnt apply to the product
    # Find the columns that are all empty and remove that index from the data and header
    header = [i for i in data[0].keys()]
    empties = []
    for x in range(len(data[0])):
        col = set([i[x] for i in data])
        if not any(col):
            empties.append(x)
    data = [list(i) for i in data]
    for x in empties[::-1]:
        header.pop(x)
        for xx, i in enumerate(data):
            data[xx].pop(x)
    data.insert(0, header)
    writerows(wb, sheet, data)

    # Installation Level
    sheet = wb.add_worksheet("Installations")
    query = f"""
    select
    account_name as "Account",
    close_date as "Next Renewal",
    renewal_qt as "Renewal Qt",
    forecast as "Renwewal Forecast",
    opp_count as "Renewal Opps",
    tier as "Tier",
    csm as "CSM",
    csm_manager as "CSM Manager",
    cse as "CSE",
    arr as "ARR",
    sub_product_arr as "ARR(Sub)",
    opp_acv as "Product ACV",
    csm_score as "CSM Score",
    gs_score as "GS Score",
    csm_comments as "CSM Comments",
    adoption_comments as "Adoption Comments",
    last_timeline as "Latest CSE Activity",
    product_usage_analytics as "Last CUA",
    tech_assessment as "Last TA",
    csa_whiteboarding as "Last WB",
    licenses_purchased as "Licenses",
    le as "LE Count",
    le_perc as "LE Perc",
    me as "ME Count",
    me_perc as "ME Perc",
    he as "HE Count",
    he_perc as "HE Perc",
    normalized_host_count as "Normalized Endpoints",
    deployment as "Deployment",
    last_contact as "Last Contact",
    air_gapped as "Connected",
    inst_id as "Installation ID",
    sid as "SID",
    acct_id as "Account ID"
    from inst_summary
    where product = '{prod}'
    order by account_name;
    """
    data = db.execute_dict(query)

    # Clean up data that doesnt apply to the product
    # Find the columns that are all empty and remove that index from the data and header
    empties = []
    for x in range(len(data[0])):
        col = set([i[x] for i in data])
        if not any(col):
            empties.append(x)
    header = [i for i in data[0].keys()]
    data = [list(i) for i in data]
    for x in empties[::-1]:
        header.pop(x)
        for xx, i in enumerate(data):
            data[xx].pop(x)
    data.insert(0, header)
    writerows(wb, sheet, data)

    wb.close()

if __name__ == "__main__":
    #for prod in ("Cb Protection", "Cb Response", "Cb Response Cloud"):
    table_creations()
    rd = report_data()
    #rd.get_activity()
    rd.get_installation_info()
    rd.get_account_info()
    rd.get_opportunity_info()
    rd.get_subscription_info()
    rd.get_cta_info()
    rd.renewal_quarter()
    rd.deployment_percentage()
    rd.enforcement_levels()
    rd.air_gapped()
    rd.get_s3()
    rd.product_family()
    for prod in ["Cb Response Cloud"]:
        db = sqlite_db("onprem_products.db")
        acct_data = create_acct_master(db, prod)
        inst_data = create_inst_master(db, prod)
        write_report(db, prod)
